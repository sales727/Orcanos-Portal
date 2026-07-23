from io import BytesIO

from openpyxl import load_workbook


def _read_sheet(ws):
    """Read a worksheet and return (headers, all_rows, preview_rows, row_count)."""
    hdrs = []
    for cell in ws[1]:
        hdrs.append(cell.value if cell.value is not None else "")

    all_rows = []
    preview = []
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_values = [cell.value for cell in row]
        if all(v is None for v in row_values):
            continue
        row_dict = {}
        for col_idx, cell in enumerate(row):
            header = hdrs[col_idx] if col_idx < len(hdrs) else f"Column {col_idx + 1}"
            row_dict[header] = cell.value
        count += 1
        row_dict["_originalRowNumber"] = count
        all_rows.append(row_dict)
        if len(preview) < 5:
            preview.append(row_dict)

    return hdrs, all_rows, preview, count


def read_workbook_bytes(file_bytes: bytes, main_sheet_name: str = "", steps_sheet_name: str = ""):
    """
    Parse an uploaded .xlsx file. Mirrors the original Flask /api/upload behaviour:
    reads the requested main sheet (or the active sheet if none given yet), plus an
    optional steps sheet, and returns everything the frontend needs to render a
    preview and, on a later call, the full dataset for mapping/import.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet_names = wb.sheetnames

    if main_sheet_name and main_sheet_name in sheet_names:
        ws_main = wb[main_sheet_name]
    else:
        ws_main = wb.active

    main_headers, main_rows, main_preview, main_count = _read_sheet(ws_main)

    steps_headers = None
    steps_data = None
    if steps_sheet_name and steps_sheet_name not in ("None", "") and steps_sheet_name in sheet_names:
        ws_steps = wb[steps_sheet_name]
        steps_headers, steps_data, _, _ = _read_sheet(ws_steps)

    wb.close()

    result = {
        "headers": main_headers,
        "preview": main_preview,
        "data": main_rows,
        "totalRows": main_count,
        "sheetNames": sheet_names,
    }
    if steps_headers is not None:
        result["stepsHeaders"] = steps_headers
    if steps_data is not None:
        result["stepsData"] = steps_data

    return result
