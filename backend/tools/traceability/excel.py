from io import BytesIO

from openpyxl import load_workbook


def parse_traceability_rows(
    file_bytes: bytes,
    sheet_name: str,
    source_col: str,
    target_col: str,
    relation_col: str,
):
    """Parse an uploaded .xlsx traceability sheet into row dicts.

    Mirrors the importer tool's openpyxl-based parsing. A row is marked
    invalid (and skipped by the caller) when Source Key or Target Key is
    blank — Relation Type is only required for the "add" action, checked
    by the caller since this parser doesn't know the action.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet_names = wb.sheetnames

    ws = wb[sheet_name] if sheet_name and sheet_name in sheet_names else wb.active

    headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
    missing = [c for c in (source_col, target_col, relation_col) if c and c not in headers]

    rows = []
    row_number = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        values = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
        if all(v is None for v in values.values()):
            continue
        row_number += 1

        source_key = values.get(source_col)
        target_key = values.get(target_col)
        relation_type = values.get(relation_col)

        rows.append({
            "row": row_number,
            "sourceKey": str(source_key) if source_key is not None else "",
            "targetKey": str(target_key) if target_key is not None else "",
            "relationType": str(relation_type) if relation_type is not None else "",
            "valid": source_key is not None and target_key is not None,
        })

    wb.close()

    return {
        "sheetNames": sheet_names,
        "headers": headers,
        "missingColumns": missing,
        "rows": rows,
        "totalRows": len(rows),
        "validRows": sum(1 for r in rows if r["valid"]),
    }
